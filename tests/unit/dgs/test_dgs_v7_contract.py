from pathlib import Path

from openpyxl import load_workbook

from gis2dgs.config import load_dgs_schema, load_project_config
from gis2dgs.dgs import DgsMapper, DgsWriter, inspect_excel_template
from gis2dgs.domain import Bus, Line, NetworkModel, Source
from gis2dgs.domain.identifiers import BusId, FeederId, LineId, SourceId
from gis2dgs.electrical import ElectricalLibrary, LineType
from gis2dgs.pipeline import run_conversion
from gis2dgs.powerfactory import PowerFactoryMapper, PowerFactoryMappingPolicy


ROOT = Path(__file__).resolve().parents[3]


def test_minimal_example_emits_dgs_v7_profile() -> None:
    project_path = ROOT / "examples" / "minimal" / "project.yaml"
    result = run_conversion(load_project_config(project_path))
    assert result.output_dgs.exists()

    inspection = inspect_excel_template(result.output_dgs)
    assert inspection.dgs_format_version in {"7.0", "7"}

    workbook = load_workbook(result.output_dgs, data_only=False)
    try:
        general_headers = [
            str(cell.value)
            for cell in workbook["General"][1]
            if cell.value is not None
        ]
        assert any(h.startswith("FID") for h in general_headers)

        line_headers = [
            str(cell.value)
            for cell in workbook["ElmLne"][1]
            if cell.value is not None
        ]
        assert any(h.startswith("FID") for h in line_headers)
        assert any(h.startswith("OP") for h in line_headers)
        assert not any(h.startswith("bus1") for h in line_headers)
        assert not any(h.startswith("bus2") for h in line_headers)
        assert workbook["ElmLne"].cell(2, line_headers.index("OP(a:1)") + 1).value == "C"

        cubic_headers = [
            str(cell.value)
            for cell in workbook["StaCubic"][1]
            if cell.value is not None
        ]
        assert any(h.startswith("obj_id") for h in cubic_headers)
        assert workbook["StaCubic"].max_row >= 2
    finally:
        workbook.close()


def test_v7_schema_serializes_feeders_graphics_and_stacubic_only(tmp_path: Path) -> None:
    network = NetworkModel()
    network.add_bus(Bus(BusId("F1"), "0101", 10.0, feeder_id=FeederId("0101")))
    network.add_bus(Bus(BusId("L1"), "L1", 10.0, feeder_id=FeederId("0101")))
    network.add_line(Line(LineId("T1"), "T1", BusId("F1"), BusId("L1"), 0.1, 10.0, "LT"))
    network.add_source(Source(SourceId("F1"), "0101", BusId("F1"), 10.0))
    library = ElectricalLibrary.from_types(
        line_types=[LineType("LT", "LT", 10.0, 0.2, 0.1, 200.0)]
    )
    model = PowerFactoryMapper(
        PowerFactoryMappingPolicy(require_type_references=False)
    ).map(network, library)

    assert model.find_by_class("ElmFeeder")
    assert model.find_by_class("IntGrfnet")
    assert model.find_by_class("StaCubic")

    schema = load_dgs_schema(ROOT / "examples" / "minimal" / "config" / "dgs_mapping.yaml")
    document = DgsMapper(schema).map_powerfactory_model(model)
    out = tmp_path / "v7.xlsx"
    DgsWriter(schema).write(document, out)

    workbook = load_workbook(out, data_only=False)
    try:
        assert "ElmFeeder" in workbook.sheetnames
        assert "IntGrfnet" in workbook.sheetnames
        assert "IntGrf" in workbook.sheetnames
        assert workbook["ElmFeeder"].max_row >= 2
        assert workbook["IntGrfnet"].max_row >= 2
        line_headers = [
            str(cell.value)
            for cell in workbook["ElmLne"][1]
            if cell.value is not None
        ]
        assert "bus1(p)" not in line_headers
    finally:
        workbook.close()
