from pathlib import Path

from openpyxl import Workbook, load_workbook

from gis2dgs.dgs import (
    DgsClassMapping,
    DgsDocument,
    DgsIdentityMapping,
    DgsMappingProfile,
    DgsRow,
    DgsTable,
    DgsWriter,
)


def _template(path: Path) -> None:
    workbook = Workbook()
    metadata = workbook.active
    metadata.title = "General"
    metadata["A1"] = "KEEP"
    sheet = workbook.create_sheet("ElmTerm")
    sheet.append(["FID", "loc_name", "parent", "uknom"])
    sheet.append(["OLD", "old row", None, 1.0])
    workbook.save(path)
    workbook.close()


def _profile(template: Path | None, *, allow_create: bool = False) -> DgsMappingProfile:
    return DgsMappingProfile(
        configured=True,
        template_path=template,
        allow_create_without_template=allow_create,
        classes={
            "ElmTerm": DgsClassMapping(
                table="ElmTerm",
                identity=DgsIdentityMapping("FID", "loc_name", "parent"),
                required_columns=("uknom",),
            )
        },
    )


def _document() -> DgsDocument:
    document = DgsDocument()
    table = DgsTable("ElmTerm", ("FID", "loc_name", "parent", "uknom"))
    table.add(DgsRow("B1", {"FID": "B1", "loc_name": "Bus 1", "parent": None, "uknom": 10.0}))
    document.add_table(table)
    return document


def test_writer_clones_template_preserves_other_sheets_and_replaces_rows(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    target = tmp_path / "network.xlsx"
    _template(template)

    DgsWriter(_profile(template)).write(_document(), target)

    workbook = load_workbook(target, data_only=False)
    try:
        assert workbook["General"]["A1"].value == "KEEP"
        sheet = workbook["ElmTerm"]
        assert sheet["A2"].value == "B1"
        assert sheet["B2"].value == "Bus 1"
        assert sheet["D2"].value == 10.0
        assert sheet["A3"].value is None
    finally:
        workbook.close()


def test_writer_appends_missing_schema_columns(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ElmTerm"
    sheet.append(["FID", "loc_name", "parent"])
    workbook.save(template)
    workbook.close()

    DgsWriter(_profile(template)).write(_document(), tmp_path / "out.xlsx")
    workbook = load_workbook(tmp_path / "out.xlsx")
    try:
        headers = [cell.value for cell in workbook["ElmTerm"][1]]
        assert "uknom" in headers
        assert workbook["ElmTerm"]["A2"].value == "B1"
    finally:
        workbook.close()


def test_writer_can_create_workbook_only_when_explicitly_enabled(tmp_path: Path) -> None:
    output = tmp_path / "created.xlsx"
    DgsWriter(_profile(None, allow_create=True)).write(_document(), output)
    workbook = load_workbook(output)
    try:
        assert workbook["ElmTerm"]["A1"].value == "FID"
        assert workbook["ElmTerm"]["A2"].value == "B1"
    finally:
        workbook.close()
