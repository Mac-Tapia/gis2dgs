from pathlib import Path

from openpyxl import load_workbook

from gis2dgs.config import load_project_config
from gis2dgs.dgs import inspect_excel_template
from gis2dgs.pipeline import run_conversion


def test_minimal_example_runs_end_to_end() -> None:
    root = Path(__file__).parents[2]
    project_path = root / "examples" / "minimal" / "project.yaml"
    result = run_conversion(load_project_config(project_path))

    assert result.output_dgs.exists()
    assert result.buses == 2
    assert result.lines == 1
    inspection = inspect_excel_template(result.output_dgs)
    assert inspection.dgs_format_version in {"7.0", "7"}
    assert any(sheet.sheet == "StaCubic" for sheet in inspection.sheets)
    line_sheet = next(sheet for sheet in inspection.sheets if sheet.sheet == "ElmLne")
    assert not any("bus1" in column for column in line_sheet.columns)
    assert any(column.startswith("FID") for column in line_sheet.columns)
    assert any(column.startswith("OP") for column in line_sheet.columns)
    workbook = load_workbook(result.output_dgs, data_only=False)
    try:
        headers = [
            str(cell.value)
            for cell in workbook["ElmLne"][1]
            if cell.value is not None
        ]
        op_index = headers.index(next(name for name in headers if name.startswith("OP")))
        assert workbook["ElmLne"].cell(2, op_index + 1).value == "C"
        assert workbook["StaCubic"].max_row >= 2
    finally:
        workbook.close()
