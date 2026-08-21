from pathlib import Path

from openpyxl import Workbook, load_workbook

from gis2dgs.dgs import (
    DgsClassMapping,
    DgsIdentityMapping,
    DgsMapper,
    DgsMappingProfile,
    DgsReferenceMapping,
    DgsValueMapping,
    DgsWriter,
)
from gis2dgs.domain import Bus, Line, NetworkModel, Source
from gis2dgs.domain.identifiers import BusId, LineId, SourceId
from gis2dgs.electrical import ElectricalLibrary, LineType
from gis2dgs.powerfactory import PowerFactoryMapper, PowerFactoryMappingPolicy


def _template(path: Path) -> None:
    workbook = Workbook()
    general = workbook.active
    general.title = "General"
    general.append(["REFERENCE", "DO_NOT_DELETE"])
    sheets = {
        "ElmNet": ["FID", "loc_name", "parent"],
        "ElmTerm": ["FID", "loc_name", "parent", "uknom"],
        "StaCubic": ["FID", "loc_name", "parent", "obj_id"],
        "TypLne": ["FID", "loc_name", "parent", "uline", "rline", "xline", "cline", "sline"],
        "ElmLne": ["FID", "loc_name", "parent", "dline", "typ_id", "bus1", "bus2", "outserv"],
        "ElmXnet": ["FID", "loc_name", "parent", "bus1", "outserv", "uknom"],
    }
    for name, header in sheets.items():
        sheet = workbook.create_sheet(name)
        sheet.append(header)
    workbook.save(path)
    workbook.close()


def _profile(template: Path) -> DgsMappingProfile:
    common = lambda table: DgsIdentityMapping("FID", "loc_name", "parent")
    return DgsMappingProfile(
        configured=True,
        template_path=template,
        strict_unmapped_attributes=False,
        classes={
            "ElmNet": DgsClassMapping("ElmNet", common("ElmNet")),
            "ElmTerm": DgsClassMapping(
                "ElmTerm",
                common("ElmTerm"),
                attributes={"nominal_voltage_kv": DgsValueMapping("uknom")},
            ),
            "StaCubic": DgsClassMapping(
                "StaCubic",
                common("StaCubic"),
                references={"connected_element": DgsReferenceMapping("obj_id")},
            ),
            "TypLne": DgsClassMapping(
                "TypLne",
                common("TypLne"),
                attributes={
                    "nominal_voltage_kv": DgsValueMapping("uline"),
                    "r1_ohm_per_km": DgsValueMapping("rline"),
                    "x1_ohm_per_km": DgsValueMapping("xline"),
                    "c1_nf_per_km": DgsValueMapping("cline"),
                    "rated_current_a": DgsValueMapping("sline", scale=0.001),
                },
            ),
            "ElmLne": DgsClassMapping(
                "ElmLne",
                common("ElmLne"),
                attributes={
                    "length_km": DgsValueMapping("dline"),
                    "in_service": DgsValueMapping("outserv", value_map={"true": 0, "false": 1}),
                },
                references={
                    "type": DgsReferenceMapping("typ_id"),
                    "terminal_1_cubicle": DgsReferenceMapping("bus1"),
                    "terminal_2_cubicle": DgsReferenceMapping("bus2"),
                },
            ),
            "ElmXnet": DgsClassMapping(
                "ElmXnet",
                common("ElmXnet"),
                attributes={
                    "nominal_voltage_kv": DgsValueMapping("uknom"),
                    "in_service": DgsValueMapping("outserv", value_map={"true": 0, "false": 1}),
                },
                references={"cubicle": DgsReferenceMapping("bus1")},
            ),
        },
    )


def test_network_to_powerfactory_to_dgs_excel(tmp_path: Path) -> None:
    network = NetworkModel()
    network.add_bus(Bus(BusId("B1"), "B1", 10.0))
    network.add_bus(Bus(BusId("B2"), "B2", 10.0))
    network.add_line(Line(LineId("L1"), "L1", BusId("B1"), BusId("B2"), 0.5, 10.0, "LT"))
    network.add_source(Source(SourceId("GRID"), "Grid", BusId("B1"), 10.0))
    library = ElectricalLibrary.from_types(
        line_types=[LineType("LT", "LT", 10.0, 0.2, 0.1, 400.0, c1_nf_per_km=200.0)]
    )
    pf_model = PowerFactoryMapper(
        PowerFactoryMappingPolicy(
            create_feeder_graphics=False,
            create_feeder_objects=False,
            ensure_feeder_sources=False,
        )
    ).map(network, library)

    template = tmp_path / "reference.xlsx"
    output = tmp_path / "network_dgs.xlsx"
    _template(template)
    profile = _profile(template)

    document = DgsMapper(profile).map_powerfactory_model(pf_model)
    DgsWriter(profile).write(document, output)

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook["General"]["A1"].value == "REFERENCE"
        assert workbook["ElmLne"]["A2"].value.startswith("GIS2DGS:line:")
        assert workbook["ElmLne"]["D2"].value == 0.5
        assert workbook["TypLne"]["H2"].value == 0.4
        assert workbook["ElmXnet"]["A2"].value.startswith("GIS2DGS:source:")
    finally:
        workbook.close()
