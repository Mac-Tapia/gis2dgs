from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .columns import DgsColumnDefinition
from .exceptions import DgsTemplateError


@dataclass(frozen=True, slots=True)
class DgsSheetInspection:
    sheet: str
    header_row: int | None
    columns: tuple[str, ...]
    definitions: tuple[DgsColumnDefinition, ...]
    max_row: int
    max_column: int


@dataclass(frozen=True, slots=True)
class DgsTemplateInspection:
    path: Path
    dgs_format_version: str | None
    sheets: tuple[DgsSheetInspection, ...]

    def as_dict(self) -> dict[str, Any]:
        return {
            "template": str(self.path),
            "dgs_format_version": self.dgs_format_version,
            "sheets": [
                {
                    "sheet": sheet.sheet,
                    "header_row": sheet.header_row,
                    "columns": list(sheet.columns),
                    "column_definitions": [
                        {
                            "raw": definition.raw,
                            "name": definition.name,
                            "type": definition.type.value,
                            "size": definition.size,
                        }
                        for definition in sheet.definitions
                    ],
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                }
                for sheet in self.sheets
            ],
        }


def _detect_dgs_format_version(workbook) -> str | None:
    if "General" not in workbook.sheetnames:
        return None
    worksheet = workbook["General"]
    for row in worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 50), values_only=True):
        values = list(row)
        if len(values) >= 3 and str(values[1]).strip().lower() == "version":
            value = values[2]
            return None if value is None else str(value)
    return None


def inspect_excel_template(
    path: Path,
    *,
    search_rows: int = 15,
    minimum_header_cells: int = 2,
) -> DgsTemplateInspection:
    """Inspect an exported DGS Excel workbook without modifying it.

    The DGS format revision is read from the General sheet when available. It is
    deliberately distinct from any PowerFactory product version.
    """

    if not path.exists():
        raise DgsTemplateError(f"DGS template does not exist: {path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise DgsTemplateError("Only .xlsx/.xlsm DGS templates are supported.")

    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets: list[DgsSheetInspection] = []
    try:
        format_version = _detect_dgs_format_version(workbook)
        for worksheet in workbook.worksheets:
            header_row: int | None = None
            columns: tuple[str, ...] = ()
            limit = min(search_rows, worksheet.max_row)
            for row_index in range(1, limit + 1):
                values = [
                    worksheet.cell(row_index, col).value
                    for col in range(1, worksheet.max_column + 1)
                ]
                candidates = tuple(
                    str(value).strip()
                    for value in values
                    if value is not None and str(value).strip()
                )
                if len(candidates) >= minimum_header_cells:
                    header_row = row_index
                    columns = candidates
                    break
            sheets.append(
                DgsSheetInspection(
                    sheet=worksheet.title,
                    header_row=header_row,
                    columns=columns,
                    definitions=tuple(DgsColumnDefinition.parse(value) for value in columns),
                    max_row=worksheet.max_row,
                    max_column=worksheet.max_column,
                )
            )
    finally:
        workbook.close()

    return DgsTemplateInspection(
        path=path,
        dgs_format_version=format_version,
        sheets=tuple(sheets),
    )
