from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .exceptions import DgsSchemaNotConfiguredError, DgsTemplateError
from .models import DgsDocument
from .schema import DgsClassMapping, DgsFormat, DgsSchema


class DgsWriter:
    """Serialize a DgsDocument using a schema-driven DGS Excel template.

    The writer deliberately clones a reference workbook when available. This
    preserves template metadata sheets and columns that are not owned by
    gis2dgs instead of attempting to recreate undocumented details.
    """

    def __init__(self, schema: DgsSchema | None = None) -> None:
        self.schema = schema

    def write(self, document: DgsDocument, destination: Path) -> Path:
        schema = self._require_schema()
        if schema.format != DgsFormat.EXCEL:
            raise DgsTemplateError(f"Unsupported DGS format: {schema.format}")
        if destination.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise DgsTemplateError("Phase 8 Excel writer requires .xlsx or .xlsm output.")

        workbook = self._open_workbook(schema)
        try:
            self._ensure_general_sheet(workbook, schema)
            if not schema.preserve_unmapped_sheets:
                self._remove_unmapped_sheets(workbook, document, preserve_general=True)
            self._write_document(workbook, document, schema)
            destination.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(destination)
        finally:
            workbook.close()
        return destination

    def _require_schema(self) -> DgsSchema:
        if self.schema is None:
            raise DgsSchemaNotConfiguredError(
                "DGS writer requires a configured DGS schema."
            )
        self.schema.require_configured()
        return self.schema

    @staticmethod
    def _open_workbook(schema: DgsSchema):
        template = schema.template_path
        if template is not None and template.exists():
            keep_vba = template.suffix.lower() == ".xlsm"
            return load_workbook(template, keep_vba=keep_vba)
        if not schema.allow_create_without_template:
            raise DgsTemplateError(
                "Configured DGS template is missing. Provide a DGS reference workbook "
                "or explicitly enable template-free creation."
            )
        workbook = Workbook()
        default = workbook.active
        workbook.remove(default)
        return workbook


    @staticmethod
    def _ensure_general_sheet(workbook, schema: DgsSchema) -> None:
        if schema.dgs_format_version is None:
            return
        id_header = "FID(a:40)"
        if "General" not in workbook.sheetnames:
            worksheet = workbook.create_sheet(title="General", index=0)
            worksheet.append([id_header, "Descr(a:40)", "Val(a:40)"])
            worksheet.append(["1", "Version", schema.dgs_format_version])
            return
        worksheet = workbook["General"]
        found = False
        for row in range(1, max(worksheet.max_row, 2) + 1):
            raw = worksheet.cell(row, 2).value
            if raw is not None and str(raw).strip().lower() == "version":
                worksheet.cell(row, 3, value=schema.dgs_format_version)
                found = True
                break
        if not found:
            if worksheet.max_row < 1 or worksheet.cell(1, 1).value is None:
                worksheet.append([id_header, "Descr(a:40)", "Val(a:40)"])
            worksheet.append(
                [str(worksheet.max_row + 1), "Version", schema.dgs_format_version]
            )

    def _write_document(
        self,
        workbook,
        document: DgsDocument,
        schema: DgsSchema,
    ) -> None:
        table_mappings = self._table_mappings(schema)
        for table_name, table in document.tables.items():
            worksheet = self._worksheet(workbook, table_name, schema)
            mappings = table_mappings.get(table_name, ())
            header_row = self._header_row(mappings, schema)
            data_start_row = self._data_start_row(mappings, schema, header_row)

            if worksheet.max_row < header_row or not self._header_values(worksheet, header_row):
                self._create_header(worksheet, header_row, table.columns)

            column_index = self._column_index(worksheet, header_row)
            missing = set(table.columns).difference(column_index)
            if missing:
                self._append_missing_headers(worksheet, header_row, missing)
                column_index = self._column_index(worksheet, header_row)

            if schema.clear_existing_rows:
                self._clear_rows(worksheet, data_start_row, tuple(column_index.values()))

            for offset, row in enumerate(table.rows):
                target_row = data_start_row + offset
                for column, value in row.values.items():
                    worksheet.cell(target_row, column_index[column], value=value)

    @staticmethod
    def _table_mappings(schema: DgsSchema) -> dict[str, tuple[DgsClassMapping, ...]]:
        grouped: dict[str, list[DgsClassMapping]] = {}
        for mapping in schema.classes.values():
            grouped.setdefault(mapping.table, []).append(mapping)
        return {name: tuple(values) for name, values in grouped.items()}

    @staticmethod
    def _header_row(
        mappings: tuple[DgsClassMapping, ...],
        schema: DgsSchema,
    ) -> int:
        rows = {mapping.header_row for mapping in mappings if mapping.header_row is not None}
        if len(rows) > 1:
            raise DgsTemplateError("Classes sharing one DGS table use conflicting header rows.")
        return rows.pop() if rows else schema.default_header_row

    @staticmethod
    def _data_start_row(
        mappings: tuple[DgsClassMapping, ...],
        schema: DgsSchema,
        header_row: int,
    ) -> int:
        rows = {
            mapping.data_start_row
            for mapping in mappings
            if mapping.data_start_row is not None
        }
        if len(rows) > 1:
            raise DgsTemplateError(
                "Classes sharing one DGS table use conflicting data-start rows."
            )
        row = rows.pop() if rows else schema.default_data_start_row
        if row <= header_row:
            raise DgsTemplateError("DGS data_start_row must be after the header row.")
        return row

    @staticmethod
    def _worksheet(workbook, table_name: str, schema: DgsSchema) -> Worksheet:
        if table_name in workbook.sheetnames:
            return workbook[table_name]
        if schema.format != DgsFormat.EXCEL:
            raise DgsTemplateError(f"Unsupported DGS format: {schema.format}")
        return workbook.create_sheet(title=table_name)

    @staticmethod
    def _header_values(worksheet: Worksheet, header_row: int) -> tuple[str, ...]:
        values: list[str] = []
        for column in range(1, worksheet.max_column + 1):
            value = worksheet.cell(header_row, column).value
            if value is not None and str(value).strip():
                values.append(str(value).strip())
        return tuple(values)

    @staticmethod
    def _column_index(worksheet: Worksheet, header_row: int) -> dict[str, int]:
        index: dict[str, int] = {}
        for column in range(1, worksheet.max_column + 1):
            value = worksheet.cell(header_row, column).value
            if value is None:
                continue
            name = str(value).strip()
            if not name:
                continue
            if name in index:
                raise DgsTemplateError(
                    f"DGS template sheet {worksheet.title!r} has duplicate header {name!r}."
                )
            index[name] = column
        return index

    @staticmethod
    def _create_header(worksheet: Worksheet, header_row: int, columns: tuple[str, ...]) -> None:
        for index, column in enumerate(columns, start=1):
            worksheet.cell(header_row, index, value=column)

    @staticmethod
    def _append_missing_headers(
        worksheet: Worksheet,
        header_row: int,
        missing: set[str],
    ) -> None:
        next_column = max(worksheet.max_column, 0) + 1
        for column in sorted(missing):
            worksheet.cell(header_row, next_column, value=column)
            next_column += 1

    @staticmethod
    def _clear_rows(
        worksheet: Worksheet,
        data_start_row: int,
        mapped_columns: tuple[int, ...],
    ) -> None:
        if worksheet.max_row < data_start_row:
            return
        for row_index in range(data_start_row, worksheet.max_row + 1):
            for column_index in mapped_columns:
                worksheet.cell(row_index, column_index).value = None

    @staticmethod
    def _remove_unmapped_sheets(
        workbook,
        document: DgsDocument,
        *,
        preserve_general: bool = False,
    ) -> None:
        keep = set(document.tables)
        if preserve_general:
            keep.add("General")
        for sheet_name in tuple(workbook.sheetnames):
            if sheet_name not in keep:
                workbook.remove(workbook[sheet_name])

    @staticmethod
    def copy_cell_style(source, target) -> None:
        """Small utility retained for future template row-style propagation."""
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
